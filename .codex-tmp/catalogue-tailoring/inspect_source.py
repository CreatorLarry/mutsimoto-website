from pathlib import Path
from openpyxl import load_workbook

path = Path(r"C:\Users\ADMIN\Downloads\FINAL 1 SERIES CATALOGUE.xlsx")
workbook = load_workbook(path, read_only=True, data_only=False)

print(f"sheets={workbook.sheetnames}")
for sheet in workbook.worksheets:
    print(f"\nSHEET {sheet.title!r} dimension={sheet.calculate_dimension(force=True)} max_row={sheet.max_row} max_column={sheet.max_column}")
    shown = 0
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 80), max_col=min(sheet.max_column, 60), values_only=True), start=1):
        populated = [(index + 1, value) for index, value in enumerate(row) if value not in (None, "")]
        if populated:
            print(f"row {row_index}: {populated}")
            shown += 1
        if shown >= 35:
            break
