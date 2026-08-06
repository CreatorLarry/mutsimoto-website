import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


SOURCE_PATH = Path(r"C:\Users\ADMIN\Downloads\FINAL 1 SERIES CATALOGUE.xlsx")
OUTPUT_DIR = Path(r"C:\OneDrive_CreatorLarry\OneDrive\Documents\mutsimoto-website\outputs\catalogue-tailoring-20260805")
IMAGES_DIR = OUTPUT_DIR / "images"
JSON_PATH = OUTPUT_DIR / "catalogue-data.json"

EMPTY_MARKERS = {"", "n/a", "na", "#n/a", "none", "nil", "-"}
INDUSTRIAL_MAKES = {
    "ATLAS COPCO",
    "CUMMINS",
    "JIANG DONG",
    "JUAG DONG",
    "KAMA",
    "KIRLOSKAR",
    "LISTER PETTER",
    "LOCOMOTIVE",
    "PERKINS",
    "ROBIN",
    "TAURUS",
    "WORK ORDER",
    "YANMAR",
}
BOTH_MAKES = {"LUCAS", "RACOR"}
INDUSTRIAL_TERMS = {
    "COMPRESSOR",
    "DIESEL ENGINE",
    "DOZER",
    "EXCAVATOR",
    "GENERATOR",
    "INDUSTRIAL",
    "LOADER",
    "LOCOMOTIVE",
    "TRACTOR",
}
ACRONYMS = {"CAV", "ERF", "FAW", "FOTON", "FUSO", "HINO", "ISUZU", "OEM", "TATA", "UD"}


def text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def usable(value):
    cleaned = text(value)
    return "" if cleaned.lower() in EMPTY_MARKERS else cleaned


def normalized(value):
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def display_name(value):
    words = []
    for word in usable(value).split():
        upper = word.upper()
        words.append(upper if upper in ACRONYMS else word.capitalize())
    return " ".join(words) or "Mutsimoto"


def part_filename(value):
    return re.sub(r"[^A-Za-z0-9]+", "-", value).strip("-").upper()


def application_type(make, description, model):
    upper_make = make.upper()
    combined = f"{description} {model}".upper()
    if upper_make in BOTH_MAKES:
        return "Both"
    if upper_make in INDUSTRIAL_MAKES or any(term in combined for term in INDUSTRIAL_TERMS):
        return "Industrial"
    return "Automotive"


def equipment_details(make, model, description):
    combined = f"{make} {model} {description}".upper()
    if "GENERATOR" in combined:
        return "Generator", "Power Generation"
    if "TRACTOR" in combined:
        return "Agricultural Machinery", "Agriculture"
    if any(term in combined for term in ("EXCAVATOR", "DOZER", "LOADER")):
        return "Construction Equipment", "Construction"
    if "LOCOMOTIVE" in combined:
        return "Industrial Equipment", "Rail"
    return "Industrial Equipment", "Industrial"


workbook = load_workbook(SOURCE_PATH, read_only=False, data_only=True)
sheet = workbook["FINAL 1 SERIES CATALOGUE"]

products = []
specifications = []
references = []
vehicle_applications = []
equipment_applications = []
source_notes = []
part_by_row = {}
make_by_part = {}

for row_number in range(3, 1001):
    values = [sheet.cell(row=row_number, column=column).value for column in range(1, 35)]
    part_number = usable(values[1])
    if not part_number:
        continue

    make = usable(values[6]) or "Unspecified"
    make_display = display_name(make)
    description = usable(values[7])
    model = usable(values[8]) or "Unspecified model"
    source_category = usable(values[21]) or "Fuel Filter"
    product_kind = "Fuel Water Separator" if "water separator" in source_category.lower() else "Fuel Filter"
    broad_application = application_type(make, description, model)
    name = f"{make_display} {product_kind} {part_number}"
    short_description = description[:320] if len(description) >= 10 else ""
    full_description = description[:5000] if len(description) >= 20 else ""

    products.append([
        part_number,
        name[:160],
        "Fuel Elements",
        broad_application,
        "Contact for availability",
        "No",
        short_description,
        full_description,
        f"{part_number} {product_kind} | Mutsimoto"[:160],
        short_description,
        "",
    ])
    part_by_row[row_number] = part_number
    make_by_part[part_number] = make_display

    specification_values = [
        ("Height", values[9], "mm"),
        ("Outer Diameter", values[10], "mm"),
        ("Inner Diameter 1", values[11], "mm"),
        ("Inner Diameter 2", values[12], "mm"),
        ("Gaskets Applicable", values[24], ""),
        ("Use With", values[5], ""),
        ("Additional Usage", values[22], ""),
        ("Source Filter Type", source_category, ""),
    ]
    order = 1
    for label, raw_value, unit in specification_values:
        clean_value = usable(raw_value)
        if clean_value:
            numeric_value = raw_value if isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool) else clean_value
            specifications.append([part_number, label, numeric_value, unit, order])
            order += 1

    seen_references = set()

    def add_reference(reference_type, manufacturer, raw_value):
        reference_number = usable(raw_value)
        if not reference_number:
            return
        key = (reference_type.lower(), normalized(reference_number))
        if not key[1] or key in seen_references:
            return
        seen_references.add(key)
        references.append([part_number, reference_type, manufacturer, reference_number])

    for column_index in [2, 3, 4, 25, 26, 27, 28, 29, 30, 31]:
        add_reference("OEM", make_display, values[column_index])
    add_reference("Alternative", "Mutsimoto", values[13])
    for column_index, manufacturer in [
        (14, "Baldwin"),
        (15, "Fleetguard"),
        (16, "Donaldson"),
        (17, "JS"),
        (18, "Sakura"),
        (19, "Hengst"),
    ]:
        add_reference("Competitor", manufacturer, values[column_index])

    notes_parts = []
    if usable(values[22]):
        notes_parts.append(f"Additional usage: {usable(values[22])}")
    if usable(values[5]):
        notes_parts.append(f"Use with: {usable(values[5])}")
    application_notes = "; ".join(notes_parts)

    if broad_application in {"Automotive", "Both"}:
        vehicle_applications.append([
            part_number,
            make_display,
            model[:160],
            make_display,
            "",
            "",
            "",
            application_notes,
        ])
    if broad_application in {"Industrial", "Both"}:
        equipment_type, industry = equipment_details(make, model, description)
        equipment_applications.append([
            part_number,
            equipment_type,
            industry,
            make_display,
            model[:160],
            make_display,
            "",
            application_notes,
        ])

    source_notes.append([
        part_number,
        row_number,
        usable(values[0]),
        usable(values[23]),
        usable(values[32]),
        source_category,
    ])

IMAGES_DIR.mkdir(parents=True, exist_ok=True)
images = []
image_counts = defaultdict(int)
anchored_images = []
for image in getattr(sheet, "_images", []):
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        continue
    row_number = marker.row + 1
    part_number = part_by_row.get(row_number)
    if not part_number:
        continue
    anchored_images.append((row_number, marker.col, image, part_number))

for _, _, image, part_number in sorted(anchored_images, key=lambda item: (item[0], item[1])):
    image_counts[part_number] += 1
    sequence = image_counts[part_number]
    extension = "jpg" if str(getattr(image, "format", "")).lower() in {"jpg", "jpeg"} else "png"
    filename = f"{part_filename(part_number)}-{sequence}.{extension}"
    (IMAGES_DIR / filename).write_bytes(image._data())
    make_display = make_by_part.get(part_number, "Mutsimoto")
    view_text = "product photo" if sequence == 1 else f"product photo view {sequence}"
    images.append([part_number, filename, f"{part_number} {make_display} fuel filter {view_text}", sequence, "Yes" if sequence == 1 else "No"])

payload = {
    "products": products,
    "specifications": specifications,
    "references": references,
    "vehicleApplications": vehicle_applications,
    "equipmentApplications": equipment_applications,
    "images": images,
    "sourceNotes": source_notes,
    "summary": {
        "products": len(products),
        "specifications": len(specifications),
        "references": len(references),
        "vehicleApplications": len(vehicle_applications),
        "equipmentApplications": len(equipment_applications),
        "images": len(images),
    },
}
JSON_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
print(json.dumps(payload["summary"], indent=2))
