from collections import Counter
from pathlib import Path
from openpyxl import load_workbook

path = Path(r"C:\Users\ADMIN\Downloads\FINAL 1 SERIES CATALOGUE.xlsx")
workbook = load_workbook(path, read_only=True, data_only=True)
sheet = workbook["FINAL 1 SERIES CATALOGUE"]

rows = []
for row in sheet.iter_rows(min_row=3, max_row=1000, min_col=1, max_col=34, values_only=True):
    part_number = row[1]
    if part_number in (None, ""):
        continue
    rows.append(row)

print(f"products={len(rows)}")
print("subcategories:")
for value, count in Counter(str(row[21]).strip() if row[21] not in (None, "") else "<blank>" for row in rows).most_common():
    print(f"  {value}: {count}")
print("makes:")
for value, count in Counter(str(row[6]).strip() if row[6] not in (None, "") else "<blank>" for row in rows).most_common(80):
    print(f"  {value}: {count}")
print("last products:")
for row in rows[-20:]:
    print(row[0], row[1], row[6], row[8], row[21])
